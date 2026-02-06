import os
from pathlib import Path


async def parse_document(file_path: str, file_type: str) -> tuple[str, int]:
    """Parse a document and return (raw_text, page_count)."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    if file_type == "pdf":
        return _parse_pdf(path)
    elif file_type in ("docx", "doc"):
        return _parse_docx(path)
    elif file_type in ("xlsx", "xls"):
        return _parse_excel(path)
    elif file_type in ("txt", "csv"):
        return _parse_text(path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def _parse_pdf(path: Path) -> tuple[str, int]:
    try:
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            text = page.extract_text() or ""
            pages.append(text)
        return "\n\n---PAGE BREAK---\n\n".join(pages), len(reader.pages)
    except ImportError:
        raise RuntimeError("pypdf is required for PDF parsing. Install with: pip install pypdf")


def _parse_docx(path: Path) -> tuple[str, int]:
    try:
        from docx import Document

        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        estimated_pages = max(1, len(text) // 3000)
        return text, estimated_pages
    except ImportError:
        raise RuntimeError("python-docx is required for DOCX parsing. Install with: pip install python-docx")


def _parse_excel(path: Path) -> tuple[str, int]:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True)
        sheets_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    rows.append("\t".join(cells))
            if rows:
                sheets_text.append(f"=== Sheet: {sheet_name} ===\n" + "\n".join(rows))
        return "\n\n".join(sheets_text), len(wb.sheetnames)
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")


def _parse_text(path: Path) -> tuple[str, int]:
    text = path.read_text(encoding="utf-8", errors="replace")
    estimated_pages = max(1, len(text) // 3000)
    return text, estimated_pages
